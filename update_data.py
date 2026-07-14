# -*- coding: utf-8 -*-
import json
import datetime
import re
import os
import sys

# Ensure stdout handles Thai characters properly
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

def parse_data():
    csv_path = 'ระบบบันทึกข้อมูลการลาออนไลน์.csv'
    excel_path = 'ระบบบันทึกข้อมูลการลาออนไลน์.xlsx'
    
    records = []
    
    if os.path.exists(excel_path):
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet_name = 'ข้อมูลการลา'
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx == 0:
                continue
            if not any(row):
                continue
                
            name = row[1]
            if not name:
                continue
                
            name = re.sub(r'\s+', ' ', str(name).strip())
            position = str(row[2]).strip() if row[2] else ""
            
            try:
                fiscal_year = int(float(row[3])) if row[3] is not None else 2569
            except ValueError:
                fiscal_year = 2569
                
            try:
                round_no = int(float(row[4])) if row[4] is not None else 1
            except ValueError:
                round_no = 1
                
            round_detail = str(row[5]).strip() if row[5] else f"รอบ {round_no}"
            leave_type = str(row[6]).strip() if row[6] else ""
            
            # Helper to format dates
            def format_date(val):
                if isinstance(val, (datetime.datetime, datetime.date)):
                    return val.strftime('%Y-%m-%d')
                elif val:
                    val_str = str(val).strip()
                    m = re.match(r'^(\d{4}-\d{2}-\d{2})', val_str)
                    if m:
                        return m.group(1)
                    return val_str
                return ""
                
            date_from = format_date(row[7])
            date_to = format_date(row[8])
            
            try:
                days = float(row[9]) if row[9] is not None else 0.0
                if days.is_integer():
                    days = int(days)
            except ValueError:
                days = 0
                
            remark = str(row[10]).strip() if row[10] else ""
            timestamp = str(row[0]).strip() if row[0] else ""
            
            records.append({
                "timestamp": timestamp,
                "name": name,
                "position": position,
                "fiscalYear": fiscal_year,
                "round": round_no,
                "roundDetail": round_detail,
                "type": leave_type,
                "dateFrom": date_from,
                "dateTo": date_to,
                "days": days,
                "remark": remark
            })
    elif os.path.exists(csv_path):
        import csv
        with open(csv_path, mode='r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            header = next(reader)
            for row in reader:
                if not row or len(row) < 2:
                    continue
                name = row[1]
                if not name:
                    continue
                name = re.sub(r'\s+', ' ', str(name).strip())
                position = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                
                try:
                    fiscal_year = int(float(row[3])) if len(row) > 3 and row[3] else 2569
                except ValueError:
                    fiscal_year = 2569
                    
                try:
                    round_no = int(float(row[4])) if len(row) > 4 and row[4] else 1
                except ValueError:
                    round_no = 1
                    
                round_detail = str(row[5]).strip() if len(row) > 5 and row[5] else f"รอบ {round_no}"
                leave_type = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                
                date_from = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                date_to = str(row[8]).strip() if len(row) > 8 and row[8] else ""
                
                try:
                    days = float(row[9]) if len(row) > 9 and row[9] else 0.0
                    if days.is_integer():
                        days = int(days)
                except ValueError:
                    days = 0
                    
                remark = str(row[10]).strip() if len(row) > 10 and row[10] else ""
                timestamp = str(row[0]).strip() if len(row) > 0 and row[0] else ""
                
                records.append({
                    "timestamp": timestamp,
                    "name": name,
                    "position": position,
                    "fiscalYear": fiscal_year,
                    "round": round_no,
                    "roundDetail": round_detail,
                    "type": leave_type,
                    "dateFrom": date_from,
                    "dateTo": date_to,
                    "days": days,
                    "remark": remark
                })
    else:
        print(f"ไม่พบไฟล์ข้อมูล Excel หรือ CSV ในโฟลเดอร์ปัจจุบัน")
        sys.exit(1)
        
    return records

if __name__ == '__main__':
    print("กำลังเริ่มอ่านข้อมูลการลา...")
    records = parse_data()
    
    # Generate Thai formatted timestamp
    thai_months_full = [
        "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
        "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
    ]
    now = datetime.datetime.now()
    thai_year = now.year + 543
    thai_month = thai_months_full[now.month - 1]
    time_str = now.strftime('%H:%M น.')
    last_updated_str = f"{now.day} {thai_month} {thai_year} เวลา {time_str}"
    
    js_content = f"""// ข้อมูลการลาที่อัปเดตอัตโนมัติจากไฟล์ Excel
const lastUpdated = "{last_updated_str}";
const leaveRecords = {json.dumps(records, ensure_ascii=False, indent=2)};
"""
    
    output_path = 'data.js'
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(js_content)
        
    print(f"ดึงข้อมูลสำเร็จ! พบรายการลาทั้งหมด {len(records)} รายการ")
    print(f"อัปเดตไฟล์ข้อมูลเรียบร้อย: {output_path}")
